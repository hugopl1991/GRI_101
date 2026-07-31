# -*- coding: utf-8 -*-
"""
Modulo de unidades operacionais do FUCA

Link do código-fonte: https://github.com/CID-ITV/ITV_TEA_FUCA-back
"""

from fuca.functions import *
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def get_Opunit_id(opunit_file: str) -> dict:
  """
    Funcao que obtem os ids das unidades operacionais.
    
    Entrada:
        opunit_file: String do arquivo que contem os ids de unidades operacioais.
        
    Saída: Retorna um dicionario com os ids.
  """
  df = pd.read_csv(opunit_file)

  uni_op = {}

  for index, row in df.iterrows():
    estado = row['ESTADO']
    opunit_name = row['Opunit'].replace(' ','_')
    opunit_id = row['object_id_']

    if estado not in uni_op:
      uni_op[estado] = {}

    uni_op[estado][opunit_id] = opunit_name

  return uni_op

def get_Opunit_metadata(opunit_file: str) -> dict:
  """
    Funcao que obtem os metadados das unidades operacionais.
    
    Entrada:
        opunit_file: String do arquivo que contem os metadados.
        
    Saída: Retorna um dicionario com os metadados por unidade operacional.
  """

  df = pd.read_csv(opunit_file,index_col=0)

  dict_opunit = {}

  estados = list( set( list(df.index.values) ) )

  colunas = list(df)

  for estado in estados:
    df_tmp = df.loc[(df.index == estado)]
    dict_opunit[estado] = []
    for index, row in df_tmp.iterrows():
      tmp = {}
      for col in colunas:
          tmp[col] = row[col]
      dict_opunit[estado].append(tmp)
  
  return dict_opunit

def make_C_Opunit_table(list_indices: List[Tuple[int,int]],
                        list_indices_id: List[Tuple[int,int]],
                        band: np.ndarray,
                        carbon_type: str) -> dict:
  """
    Funcao que calcula o carbono total em uma unidade operacional.
    
    Entrada:
        list_indices: Lista de indices do mapa de carbono.
        list_indices_id: Lista de indices do mapa de id de Opunit.
        band: Matriz numpy contendo os valores de carbono.
        carbon_type: String contendo o tipo de carbono que esta sendo calculado (Ex: TAGB_BGB).
        
    Saída: Dicionario contendo o total de carbono e a quantidade de pixels.
  """
  list_final = list( set(list_indices_id) & set(list_indices) )

  soma = 0
  for ind in list_final:
    pixel = band[ind[0]][ind[1]]
    soma+= pixel
  band = None

  tmp = {"Qtd_npixels":len(list_indices_id)}
  tmp[carbon_type+'_total_sum'] = float("{:.5f}".format(soma))

  list_final = None

  return tmp

def make_Opunit_sheet(dict_opunit: dict, dict_carbon_OpUnit: dict, areas: List[str], sheet_names: List[str],
                      years: List[int], carbon_type: str, path_out: str, name_final: str = None):
  """
    Funcao que gera a planilha de unidades operacionais.
    
    Entrada:
        dict_opunit: Dicionario contendo os metadados das unidades operacionais.
        dict_carbon_OpUnit: Dicionario  contendo os valores de carbono por unidade operacional.
        areas: Lista de areas que foram calculadas o carbono das unidades operacionais.
        sheet_names: Lista com as abas que serao criadas na planilha final.
        years: Lista contendo os anos de interesse.
        carbon_type: String contendo o tipo de carbono que esta sendo calculado (Ex: TAGB_BGB).
        path_out: String contendo o caminho onde sera salva a planilha.
        name_final: String contendo o nome final da planilha, se houver (Ex: OpUnit_carbon.xlsx).
        
    Saída:
  """
  cols_names = ['Estado','OpUnit','Operacao_Projeto','Corredor','ID-Region','Region']

  wb = openpyxl.Workbook()
  wb.remove(wb['Sheet'])

  for c,sheet_name in enumerate(sheet_names):

    lista_final = [cols_names+years]
    for area in areas:
      for dict_op in dict_opunit[area]:
        operacao = dict_op['Operacao'].replace(' ','_')
        auditoria = dict_op['Auditoria']
        corredor = dict_op['Corredor']
        id_region = dict_op['ID_Region']
        region = dict_op['Region']

        if operacao == 'Areas_Vale':
          operacao = 'Vale'

        if operacao not in dict_carbon_OpUnit[area][years[0]]:
          continue

        tmp_area = [area,operacao,auditoria,corredor,id_region,region]
        tmp_ano = []
        for year in years:
          value = dict_carbon_OpUnit[area][year][operacao][sheet_name]
          tmp_ano.append(value)

        lista_final.append(tmp_area+tmp_ano)

    wb.create_sheet(index=c, title=sheet_name)
    ws = wb.worksheets[c]

    for i in range(len(lista_final)):
      for j in range(len(lista_final[i])):

        ws.cell(row=i+1, column=j+1).value = lista_final[i][j]

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
      if get_column_letter(col) == 'A':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)
      elif get_column_letter(col) == 'B':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)
      elif get_column_letter(col) == 'C':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)
      elif get_column_letter(col) == 'D':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
      elif get_column_letter(col) == 'E':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)
      elif get_column_letter(col) == 'F':
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
      else:
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)

    ws.column_dimensions = dim_holder

  if name_final == None:
    name_final = carbon_type+'_unidades_operacionais.xlsx'

  wb.save(path_out+name_final)