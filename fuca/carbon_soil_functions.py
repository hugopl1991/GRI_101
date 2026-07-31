# -*- coding: utf-8 -*-
"""
Modulo de carbono no solo do FUCA

Link do código-fonte: https://github.com/CID-ITV/FUCA
"""

from fuca.functions import *

def get_flu_values(xlsx_path: str, area: str) -> dict:
  """
    Função que obtem os coeficientes de FLU por classe de uso do solo.
    
    Entrada:
        xlsx_path: String do arquivo xlsx que contém os coeficientes de FLU.
        area: String que contém a área de interesse.
        
    Saída: Retorna um dicionario com os coeficientes de FLU separados por classe de uso.
  """
  import openpyxl
  from pathlib import Path

  wb_obj = openpyxl.load_workbook(xlsx_path) 

  flu_values = {}
  for row in wb_obj[area].iter_rows(min_row=2,max_row=wb_obj[area].max_row):
    if row[0].value != None:
      flu_values[row[0].value] = float(row[1].value)

  return flu_values

def make_soil_C_ref_map(path_raster_map: str, path_C_solo: str, shp_path: str,
                        flu_values: dict, area: str, year: int, tmp_path: str,
                        out_path: str, all_touched: bool = False) -> str:
  """
    Função que cria o mapa de carbono do solo de referência.
    
    Entrada:
        path_raster_map: String com o caminho do mapa de uso e cobertura do solo.
        path_C_solo: String com o caminho do mapa de carbono no solo preterito.
        shp_path: String com o caminho do shapefile da area de interesse.
        flu_values: Dicionário com os coeficientes de FLU.
        area: String que contém a área de interesse.
        year: Inteiro do ano que inicia a serie historica.
        tmp_path: String com o caminho da pasta temporaria.
        out_path: String com o caminho da pasta de saida.
        all_touched: Variavel booleana que define se o recorte sera feito com all_touched=True ou False.
        
    Saída: Retorna uma string com o caminho onde mapa foi salvo.
  """

  #recorta o mapbiomas
  path_raster_map = clip_raster(raster_path=path_raster_map,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)

  #acessa o mapbiomas
  band,_,list_indices = access_raster(raster_path=path_raster_map,list_ind=True)
  #acessa o map de carbono do solo
  #_,nodata = access_raster(raster_path=path_C_solo,list_ind=False)

  dict_classes = {}
  for ind in list_indices:
    pixel = band[ind[0]][ind[1]]
    if pixel not in dict_classes:
      dict_classes[pixel] = []
      dict_classes[pixel].append(ind)
    else:
      dict_classes[pixel].append(ind)
  band, list_indices = None, None

  vetor_pixel = []
  vetor_value = []

  band,nodata = access_raster(raster_path=path_C_solo,list_ind=False)

  for veg in dict_classes:
    for ind in dict_classes[veg]:
      pixel = band[ind[0]][ind[1]]
      if pixel == nodata or pixel == 0:
        continue
      vetor_pixel.append(ind)
      value = pixel*flu_values[veg]
      vetor_value.append(float("{:.5f}".format(value)))
  
  band = None

  raster_carbon_soil = make_raster(vetor_pixels=vetor_pixel, vetor_value=vetor_value,
                                   raster_std=path_raster_map, est_interesse=area, 
                                   ano_interesse=year,tmp_path=out_path, name='Csoil_MgCha',
                                   final_name=None,nodata=-9999,dtype='float32',descriptions='Soil Carbon in Mg-C/ha')

  vetor_pixel, vetor_value = None, None

  dict_classes = None

  return raster_carbon_soil

def make_age_ref_raster(path_raster_map: str, shp_path: str, tmp_path: str, all_touched: bool = False) -> str:
  """
    Função que cria o mapa idade do carbono no solo.
    
    Entrada:
        path_raster_map: String com o caminho do mapa de uso e cobertura do solo.
        shp_path: String com o caminho do shapefile da area de interesse.
        tmp_path: String com o caminho da pasta temporaria.
        all_touched: Variavel booleana que define se o recorte sera feito com all_touched=True ou False.
        
    Saída: Retorna uma string com o caminho onde mapa de idade foi salvo.
  """

  path_raster_map = clip_raster(raster_path=path_raster_map,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)
  
  _,_,list_indices = access_raster(raster_path=path_raster_map,list_ind=True)
  vetor_value = [20 for ind in list_indices]

  raster_age = make_raster(vetor_pixels=list_indices, vetor_value=vetor_value, 
                           raster_std=path_raster_map, est_interesse='idade', 
                           ano_interesse='solo',tmp_path=tmp_path, name='ref',
                           final_name=None,nodata=255,dtype='uint8',descriptions='Raster auxiliar de idade')

  return raster_age

def generate_c_soil_maps(area: str, start_year: int, end_year: int,
                         shp_path: str, flu_values: dict, LULC_paths: str,
                         path_C_soil_ref: str, save_zip: str, del_classes: List[int],
                         all_touched: bool = False) -> None:
    """
    Função que gera os mapas de carbono no solo para a serie historica.

    Entrada:
        area: String que contém a área de interesse.
        start_year: Inteiro do primeiro ano da serie historica.
        end_year: Inteiro do ultimo ano da serie historica.
        shp_path: String com o caminho do shapefile da area de interesse.
        flu_values: Dicionario com os valores de FLU por classe de uso.
        path_C_soil_ref: String com o caminho onde esta o mapa de carbono no solo preterito.
        del_classes: Lista com as classes cujos valores de carbono no solo devem ser zero.
        all_touched: Variavel booleana que define se o recorte sera feito com all_touched=True ou False.
        
    Saída: Os resultados sao salvos na pasta de output, nao retorna nenhuma saida de resultado.
    """
    
    tmp_path = "csoil_tmp"
    tmp_path_2 = "csoil_tmp2"
    out_path = "c_soil_output_tmp"

    tmp_path = create_dir(dir_path=tmp_path)
    tmp_path_2 = create_dir(dir_path=tmp_path_2)
    out_path = create_dir(dir_path=out_path)

    path_raster_map =  LULC_paths.format(AREA=area,YEAR=start_year)

    print('Cria mapa de C solo do primeiro ano')
    make_soil_C_ref_map(path_raster_map=path_raster_map,
                        path_C_solo=path_C_soil_ref,shp_path=shp_path,
                        area=area,flu_values=flu_values,year=start_year,
                        tmp_path=tmp_path,out_path=out_path, all_touched=all_touched)

    print('Cria mapa de idade inicial')
    raster_age = make_age_ref_raster(path_raster_map=path_raster_map,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)

    raster_carbon_eq = out_path+'raster_Csoil_MgCha_'+area+'_'+str(start_year)+'.tif'
    path_C_soil_current = out_path+'raster_Csoil_MgCha_'+area+'_'+str(start_year)+'.tif'

    for year in range(start_year+1,end_year+1):
        print("Cria mapa de C solo do ano {}".format(year))

        path_raster_map_pass = LULC_paths.format(AREA=area,YEAR=year-1)
        path_raster_map_current = LULC_paths.format(AREA=area,YEAR=year)

        path_C_soil_current = out_path+'raster_Csoil_MgCha_'+area+'_'+str(year-1)+'.tif'

        path_raster_map_pass = clip_raster(raster_path=path_raster_map_pass,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)
        path_raster_map_current = clip_raster(raster_path=path_raster_map_current,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)

        #acessa raster de idade
        band_idade,_ = access_raster(raster_path=raster_age,list_ind=False)

        #acessa o mapbiomas passado
        band_pass,_ = access_raster(raster_path=path_raster_map_pass,list_ind=False)
        #acessa o mapbiomas atual
        band,_ = access_raster(raster_path=path_raster_map_current,list_ind=False)
        #acessa o mapa de C solo atual
        _,_,list_indices = access_raster(raster_path=path_C_soil_current,list_ind=True)

        #verifica quais classes mudaram
        change_pixel = []
        vetor_value = []
        vetor_pixel = []
        for ind in list_indices:
            pixel_current = band[ind[0]][ind[1]]
            if pixel_current == 0:
                continue
            pixel_pass = band_pass[ind[0]][ind[1]]
            pixel_idade = band_idade[ind[0]][ind[1]]
            if pixel_current != pixel_pass:
                change_pixel.append(ind)
                vetor_value.append(1)
            else:
                vetor_value.append(pixel_idade+1)
            vetor_pixel.append(ind)
        band, band_pass,band_idade = None,None,None

        #atualiza raster de idade
        raster_age = make_raster(vetor_pixels=vetor_pixel, vetor_value=vetor_value, 
                                    raster_std=path_raster_map_current, est_interesse='idade', 
                                    ano_interesse='solo',tmp_path=tmp_path_2, name='ref',
                                    final_name=None,nodata=255,dtype='uint8',descriptions='Raster auxiliar de idade')
        vetor_value = None

        #retira pixels que mudaram de classe
        list_indices = list(set(vetor_pixel) - set(change_pixel))

        vetor_pixel = []
        vetor_value = []

        #add os pixels que permaceram iguais
        #acessa o map de carbono do solo ref
        band,_ = access_raster(raster_path=raster_carbon_eq,list_ind=False)

        for ind in list_indices:
            pixel = band[ind[0]][ind[1]]
            vetor_pixel.append(ind)
            vetor_value.append(pixel)
        band,list_indices = None, None

        #add os pixels que mudaram
        #acessa o map de carbono do solo atual
        band,_ = access_raster(raster_path=path_C_soil_current,list_ind=False)

        for ind in change_pixel:
            pixel = band[ind[0]][ind[1]]
            vetor_pixel.append(ind)
            vetor_value.append(pixel)
        band,change_pixel = None, None

        #cria raster de C no solo equilibrado
        raster_carbon_eq = make_raster(vetor_pixels=vetor_pixel, vetor_value=vetor_value, 
                                        raster_std=path_C_soil_current, est_interesse='C', 
                                        ano_interesse='solo',tmp_path=tmp_path_2, name='eq',
                                        final_name=None,nodata=-9999,dtype='float32',descriptions='Raster Soil Carbon Eq')

        vetor_pixel, vetor_value = None, None

        #cria o novo mapa de C no solo
        #pega os valores de flu

        #acessa raster de idade
        band_idade,_,list_indices = access_raster(raster_path=raster_age,list_ind=True)
        pixel_eq = []
        for ind in list_indices:
            pixel = band_idade[ind[0]][ind[1]]
            if pixel > 20:
                pixel_eq.append(ind)
        band_idade = None

        #retira pixels que permanecem com o mesmo valor
        list_indices = list(set(list_indices) - set(pixel_eq))

        #acessa o mapbiomas atual
        band,_ = access_raster(raster_path=path_raster_map_current,list_ind=False)

        dict_classes = {}
        for ind in list_indices:
            pixel = band[ind[0]][ind[1]]
            if pixel == 0:
                continue
            if pixel not in dict_classes:
                dict_classes[pixel] = []
                dict_classes[pixel].append(ind)
            else:
                dict_classes[pixel].append(ind)
        band, list_indices = None, None

        vetor_pixel, vetor_value = [], []

        band,_ = access_raster(raster_path=path_C_soil_ref,list_ind=False)

        for id in dict_classes:
            for ind in dict_classes[id]:
                pixel = band[ind[0]][ind[1]]
                value = pixel*flu_values[id]/20
                vetor_value.append(float("{:.5f}".format(value)))
                #print(pixel,flu_values[id],end='   ')
                #print(pixel,flu_values[id])
                #sys.exit()
        band = None
        #print()
        #print(vetor_value)
        band,_ = access_raster(raster_path=raster_carbon_eq,list_ind=False)

        count = 0
        for id in dict_classes:
            for ind in dict_classes[id]:
                pixel = band[ind[0]][ind[1]]
                vetor_value[count] -= pixel/20
                vetor_value[count] = float("{:.5f}".format(vetor_value[count]))
                count += 1
        band = None
        #print(vetor_value)
        band,_ = access_raster(raster_path=path_C_soil_current,list_ind=False)

        count = 0
        for id in dict_classes:
            for ind in dict_classes[id]:
                pixel = band[ind[0]][ind[1]]
                vetor_value[count] += pixel
                vetor_value[count] = float("{:.5f}".format(vetor_value[count]))
                count += 1
                vetor_pixel.append(ind)
        #print(vetor_value)
        #print(vetor_pixel)

        for ind in pixel_eq:
            pixel = band[ind[0]][ind[1]]
            vetor_value.append(float("{:.5f}".format(pixel)))
            vetor_pixel.append(ind)
        band, pixel_eq = None, None

        #cria raster de C no solo equilibrado
        path_C_soil_current = make_raster(vetor_pixels=vetor_pixel, vetor_value=vetor_value, 
                                        raster_std=path_C_soil_current, est_interesse=area, 
                                        ano_interesse=year,tmp_path=out_path, name='Csoil_MgCha',
                                        final_name=None,nodata=-9999,dtype='float32',descriptions='Soil Carbon in Mg-C/ha')
    ######retira classes 0

    print('Retira as classes com carbono 0')
    for year in range(start_year,end_year+1):
        print("####### Ano",year)

        path_raster_map = LULC_paths.format(AREA=area,YEAR=year)

        path_C_soil = out_path+'raster_Csoil_MgCha_'+area+'_'+str(year)+'.tif'

        path_raster_map = clip_raster(raster_path=path_raster_map,shp_path=shp_path,tmp_path=tmp_path, all_touched=all_touched)
        band_map,_,list_indices = access_raster(raster_path=path_raster_map,list_ind=True)
        dict_classes = create_raster_dict(band=band_map,list_indices=list_indices)
        band_map = None

        remove_classes = {id:dict_classes[id] for id in del_classes if id in dict_classes}
        dict_classes = None

        for id in remove_classes:
            list_indices = list( set(list_indices) - set(remove_classes[id]))

        band_soil,_ = access_raster(raster_path=path_C_soil,list_ind=False)

        vetor_pixel = []
        vetor_value = []

        for ind in list_indices:
            pixel = band_soil[ind[0]][ind[1]]
            vetor_value.append(pixel)
            vetor_pixel.append(ind)

        band_soil, list_indices = None, None

        for id in remove_classes:
            for ind in remove_classes[id]:
                vetor_value.append(0)
                vetor_pixel.append(ind)

        path_raster_carbon = make_raster(vetor_pixels=vetor_pixel, vetor_value=vetor_value, 
                                        raster_std=path_C_soil, est_interesse=area, 
                                        ano_interesse=year,tmp_path=out_path, name='Csoil_MgCha',final_name=None,
                                        nodata=-9999,dtype='float32',descriptions='Soil Carbon in Mg-C/ha')

        vetor_pixel, vetor_value = None, None

    print('Cria arquivo zip final')
    for year in range(start_year,end_year+1):

        name_file = 'raster_Csoil_MgCha_'+area+'_'+str(year)+'.tif'
        if not os.path.isfile(save_zip):
            save_zip = zip_file(path_file=out_path+name_file, path_output='', name=save_zip)
        else:  
            add_zip(path_zip=save_zip, path_file=out_path+name_file)

    del_folder(dest=tmp_path)
    del_folder(dest=tmp_path_2)
    del_folder(dest=out_path)