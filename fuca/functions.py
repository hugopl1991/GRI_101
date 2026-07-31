# -*- coding: utf-8 -*-
"""
Modulo de funcoes do FUCA

Link do código-fonte: https://github.com/CID-ITV/ITV_TEA_FUCA-back
"""
def test_function():
  print("This is a test function!")
  return "This is a test function!"

import rasterio
import pyproj
from osgeo import ogr
import numpy as np
import pandas as pd
import fiona
import rasterio.mask
from shapely.geometry import shape
import geopandas as gpd
import os
import sys
import openpyxl
from pathlib import Path
from typing import *
import zipfile

#Funcao que obtem os IDs de floresta a partir de um arquivo xlsx
def get_forest_id(xlsx_path: str,list_estados: List[str]) -> dict:
  """
    Função que obtem os IDs de floresta a partir de um arquivo xlsx.
    
    Entrada:
        xlsx_path: Arquivo xlsx que contém os valores de ID.
        list_estados: Lista de areas de interesse que devem estar no arquivo.
        
    Saída: Retorna um dicionario com as classes de vegetacao existentes em cada area.
  """

  wb_obj = openpyxl.load_workbook(xlsx_path) 

  veg_p_estado = {}
  for sheet in list_estados:
    tmp = {'natural':[],'antropica':[]}
    #print(sheet)
    for row in wb_obj[sheet].iter_rows(min_row=2,max_row=wb_obj[sheet].max_row):
      if row[0].value != None:
        if row[0].value == 'natural':
          tmp['natural'].append(row[1].value)
        else:
          tmp['antropica'].append(row[1].value)
        #print()
      veg_p_estado[sheet] = tmp

    return veg_p_estado

#Funcao que obtem os valores de AGB e coeficientes de BGB de um arquivo xlsx
def get_agb_value(xlsx_path: str,veg_p_estado: dict) -> Tuple[dict,dict,dict]:
  """
    Funcao que obtem os valores de AGB e coeficientes de BGB de um arquivo xlsx.
    
    Entrada:
        xlsx_path: Arquivo xlsx que contém os valores de AGB e BGB.
        veg_p_estado: Dicionario com as classes de vegetacao existentes em cada area.
        
    Saída:
        agb_p_estado: Retorna um dicionario com os valores de AGB de cada area separado por classe.
        agb_ibge: Retorna um dicionario com os valores de AGB auxiliar, se houver.
        bgb_p_estado: Retorna um dicionario com os coeficientes de BGB para cada classe.
  """
  #Acessa os valores de AGB com as classes de Mapbiomas e IBGE

  wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True) 

  agb_p_estado = dict.fromkeys(veg_p_estado)
  agb_ibge = dict.fromkeys(veg_p_estado)
  bgb_p_estado = {}
  for sheet in veg_p_estado:
    tmp = {'natural':{},'antropica':{}}
    tmp2 = {}

    for row in wb_obj[sheet].iter_rows(min_row=2,max_row=wb_obj[sheet].max_row):
      if row[2].value != None and row[1].value == None:
        tmp2[row[2].value] = float(row[9].value)
      if row[1].value != None:
        bgb_p_estado[row[1].value] = float(row[5].value)
        if row[1].value in veg_p_estado[sheet]['natural']:
          tmp['natural'][row[1].value] = float(row[9].value)
        else:
          tmp['antropica'][row[1].value] = float(row[9].value)
    
    agb_p_estado[sheet] = tmp
    agb_ibge[sheet] = tmp2
  
  return agb_p_estado, agb_ibge, bgb_p_estado

#Funcao que acessa o arquivo xlsx com a razao de vegetacao secundaria
def get_veg_sec_coef(xlsx_path: str, est_interesse: str) -> Tuple[float,float,float]:
  """
    Funcao que acessa o arquivo xlsx que contem a razao de vegetacao secundaria.
    
    Entrada:
        xlsx_path: Arquivo xlsx que contém a razao de vegetacao secundaria.
        est_interesse: Nome da area de interesse.
        
    Saída:
        max_agb: Float do AGB máximo da vegetação secundária.
        razao_ini: Float da razao inicial para o calculo do AGB da vegetacao secundaria.
        a1: Float do valor de biomassa para idade 1.
  """

  wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
  #sheet = wb_obj.sheetnames

  max_agb, razao_ini, a1 = None, None, None

  for row in wb_obj[est_interesse].iter_rows(min_row=1,max_row=wb_obj[est_interesse].max_row):
    if row[0].value == 'Max_AGB':
      max_agb = row[1].value
    elif row[0].value == 'Razao_Inicial':
      razao_ini = row[1].value
    elif row[0].value == 'a1':
      a1 = row[1].value
  
  if max_agb == None or razao_ini == None or a1 == None:
    print('Information missing in file '+xlsx_path)
    sys.exit()
  return max_agb, razao_ini, a1

#Funcao que acessa o arquivo de ecozonas
def access_ecozones(xlsx_path: str, est_interesse: str) -> dict:
  """
    Funcao que acessa o arquivo xlsx que contem os coeficientes de ecozonas da FAO.
    
    Entrada:
        xlsx_path: Arquivo xlsx que contém os coeficientes das ecozonas.
        est_interesse: Nome da area de interesse.
        
    Saída: Retorna um dicionario com os coeficientes das ecozonas.
  """

  wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
  #sheet = wb_obj.sheetnames

  dict_ecozones = {}
  for row in wb_obj[est_interesse].iter_rows(min_row=2,max_row=wb_obj[est_interesse].max_row):
    dict_ecozones[row[0].value] = row[2].value
  
  return dict_ecozones

#Função que irá reprojetar um raster
def convert_raster(raster_path: str, tmp_path: str, crs_std: str = 'EPSG:4326') -> str:
  """
    Funcao que reprojeta um raster.
    
    Entrada:
        raster_path: Caminho onde esta o raster a ser reprojetado.
        crs_std: Codigo de projecao, por padrao usa-se o 'EPSG:4326'.
        tmp_path: Pasta onde sera salvo o novo raster reprojetado.
        
    Saída: Retorna o caminho onde foi salvo o raster reprojetado.
  """
  #Converte raster para o projeção EPSG:4326
  from rasterio.warp import calculate_default_transform, reproject, Resampling

  dst_crs = crs_std

  with rasterio.open(raster_path) as src:
      transform, width, height = calculate_default_transform(
          src.crs, dst_crs, src.width, src.height, *src.bounds)
      kwargs = src.meta.copy()
      kwargs.update({
          'crs': dst_crs,
          'transform': transform,
          'width': width,
          'height': height
      })

      name = raster_path.replace("\\","/")
      name = name.split("/")[-1]
      with rasterio.open(tmp_path+'reproj_'+name, 'w', **kwargs) as dst:
          for i in range(1, src.count + 1):
              reproject(
                  source=rasterio.band(src, i),
                  destination=rasterio.band(dst, i),
                  src_transform=src.transform,
                  src_crs=src.crs,
                  dst_transform=transform,
                  dst_crs=dst_crs,
                  resampling=Resampling.average)
  src, dst, transform = None, None, None

  return tmp_path+'reproj_'+name

#Função que usa um shapefile para recortar um raster
def clip_raster(raster_path: str, shp_path: str = None, tmp_path: str = '', 
                column_name: str = None, column_value: Any = None, dtype: Union[str,None] = None,
                shapes: List = None, final_name: str = None, all_touched: bool = False) -> str:
  """
    Funcao que recorta um raster a partir de um shapefile.
    
    Entrada:
        raster_path: Caminho onde esta o raster a ser recortado.
        shp_path: Caminho onde está o shape que sera usado para recortar.
        tmp_path: Pasta onde sera salvo o novo raster recortado.
        column_name: String com o nome da coluna que se deseja obter os poligonos.
        column_value: Valor que a coluna deve possuir.
        dtype: String do tipo de dados do raster.
        shapes: Lista de shapes [feature geometry] que sera usada para recortar o raster.
        final_name: Nome final do arquivo raster a ser gerado.
        all_touched: Variavel booleana que indica se todos os pixels que tocarem o poligono serao recortados.
        
    Saída: Retorna o caminho onde foi salvo o raster recortado.
  """
  #Importa o shapefile do area de interesse
  if type(shapes) == type(None):
    if column_name != None and column_value != None:
        with fiona.open(shp_path, "r") as shapefile:
          shapes = [feature["geometry"] for feature in shapefile if feature['properties'][column_name] == column_value]
    else:
        with fiona.open(shp_path, "r") as shapefile:
          shapes = [feature["geometry"] for feature in shapefile]

  #Importa o raster a ser recortado
  with rasterio.open(raster_path) as src:
      out_image, out_transform = rasterio.mask.mask(src, shapes, crop=True, all_touched=all_touched)
      profile = src.profile

      if dtype == None:
        dtype = profile['dtype']

      profile.update(
        dtype=dtype,
        compress='lzw',
        height=out_image.shape[1],
        width=out_image.shape[2],
        transform=out_transform,
        driver="GTiff")

  #Cria o novo raster recortado
  
  if final_name == None:
    name = raster_path.replace("\\","/")
    name = name.split("/")[-1]
  else:
    name = final_name
    
  with rasterio.open(tmp_path+"clipped_"+name, "w", **profile) as dest:
      dest.write(out_image)
  dest, out_image, src, shapefile = None, None, None, None

  return tmp_path+"clipped_"+name

#Funcao que recorta um shapefile com outro shapefile
def clip_shp_w_shp(shp_t_clip: str, cut_shp: str, tmp_path: str) -> str:
  """
    Funcao que recorta um shapefile a partir de um shapefile.
    
    Entrada:
        shp_t_clip: Caminho onde esta o shapefile a ser recortado.
        cut_shp: Caminho onde está o shape que sera usado para recortar.
        tmp_path: Pasta onde sera salvo o novo shapefile recortado.
        
    Saída: Retorna o caminho onde foi salvo o shapefile recortado.
  """

  ## Input
  driverName = "ESRI Shapefile"
  driver = ogr.GetDriverByName(driverName)
  inDataSource = driver.Open(shp_t_clip, 0)
  inLayer = inDataSource.GetLayer()

  #print(inLayer.GetFeatureCount())
  ## Clip
  inClipSource = driver.Open(cut_shp, 0)
  inClipLayer = inClipSource.GetLayer()
  #print(inClipLayer.GetFeatureCount())

  name = shp_t_clip.replace("\\","/")
  name = name.split("/")[-1]
  outDataSource = driver.CreateDataSource(tmp_path+'clip_'+name)
  outLayer = outDataSource.CreateLayer('FINAL', geom_type=ogr.wkbMultiPolygon)

  ogr.Layer.Clip(inLayer, inClipLayer, outLayer)
  #print(outLayer.GetFeatureCount())
  inDataSource = None
  inClipSource = None
  outDataSource = None

  return tmp_path+'clip_'+name

#Funcao que acessa o shape do IBGE e gera uma lista de shapes
def get_ibge_shapes(ibge_shp: str) -> List:
  """
    Funcao que acessa o shape do IBGE e gera uma lista de shapes.
    
    Entrada:
        ibge_shp: Caminho onde esta o shapefile do IBGE.
        
    Saída:
        shapes_raster: Retorna uma lista de shapes com o ID do IBGE.
  """
  leg2_exception = [43,44,45,46,47,48,49,50]
  shapes = []
  shapes_legID = []
  shapes_legend = []
  with fiona.open(ibge_shp, "r") as shapefile:
      for feature in shapefile:
        shapes.append(feature["geometry"])
        if feature['properties']['leg2_id'] in leg2_exception:
          shapes_legID.append(feature['properties']['leg1_id'])
          shapes_legend.append(feature['properties']['legenda_1'])
        else:
          shapes_legID.append(feature['properties']['leg2_id'])
          shapes_legend.append(feature['properties']['legenda_2'])
  
  shapes_raster = []
  for i in range(len(shapes)):
    shapes_raster.append((shapes[i],shapes_legID[i]))
  
  shapefile, shapes, shapes_legID, shapes_legend = None, None, None, None
  return shapes_raster

#Funcao que acessa as ecozonas FAO e gera uma lista de shapes
def get_ecozone_shapes(eco_shp: str,col_interesse: str) -> List:
  """
    Funcao que acessa o shape de Ecozonas e gera uma lista de shapes.
    
    Entrada:
        eco_shp: Caminho onde esta o shapefile de Ecozonas.
        
    Saída:
        shapes_raster: Retorna uma lista de shapes com o ID de Ecozonas.
  """
  shapes = []
  gez_code = []
  with fiona.open(eco_shp, "r") as shapefile:
      for feature in shapefile:
        shapes.append(feature["geometry"])
        gez_code.append(feature['properties'][col_interesse])

  shapes_raster = []
  for i in range(len(shapes)):
    shapes_raster.append((shapes[i],gez_code[i]))
  
  shapefile, shapes, gez_code = None, None, None
  return shapes_raster

#Funcao que gera um raster a partir de uma lista de shapes
def shapes_t_raster(shapes_raster: List,raster_std: str, tmp_path: str, 
                    nodata: Union[int, float, str, None], dtype: str, name: str) -> str:
  """
    Funcao que gera um raster a partir de uma lista de shapes.
    
    Entrada:
        shapes_raster: Lista de shapes.
        raster_std: Caminho onde esta o raster que servira de padrao.
        tmp_path: Pasta onde sera salvo o novo shapefile recortado.
        nodata: Valor que os pixels nodata irao receber.
        dtype: Tipo de dados do raster.
        name: String com nome que o novo raster ira receber.
        
    Saída:
        Caminho do novo raster que foi gerado.
  """  
  from rasterio import features
  with rasterio.open(raster_std) as src:
      src_shape = src.shape
      transform = src.transform
      width = src.width
      height = src.height
      data_crs = src.crs
      out_meta = src.meta
  src = None

  out_meta.update({"driver": "GTiff",
                  "height": height,
                  "width": width,
                  "transform": transform})
  if dtype != None:
    out_meta.update({"dtype": dtype})
  if nodata != None:
    out_meta.update({"nodata": nodata})

  image = features.rasterize(
              ((g, v) for g, v in shapes_raster),
              out_shape=src_shape,
              transform=transform,
              all_touched=True)

  with rasterio.open(tmp_path+name+'_rasterized.tif', 'w', **out_meta) as dst:
      dst.write(image, indexes=1)
  dst, image = None, None

  return tmp_path+name+'_rasterized.tif'

#Funcao que acessa um raster e retorna a matriz de pixels e lista de indices
def access_raster(raster_path: str, list_ind: bool) -> Tuple:
  """
    Funcao que acessa um raster e retorna a matriz de pixels e lista de indices.
    
    Entrada:
        raster_path: Caminho onde esta o raster que sera acessado.
        list_ind: Variavel que verfica se sera gerado uma lista so com os pixels valorados.
        
    Saída:
        band: Matriz numpy dos pixels do raster.
        nodata: Valor nodata que existe no raster.
        list_indices: Lista com os indices dos pixels valorados.
  """ 
  from scipy.sparse import csr_matrix
  with rasterio.open(raster_path) as dataset:
    band = dataset.read(1)
    nodata = dataset.nodata
    dtype = dataset.dtypes[0]
  dataset = None

  if dtype == "float32":
    no_data = -9999
    aux = -1000
  else:
    no_data = 255
    aux = 254

  if list_ind:
    if nodata != None and nodata != 0:
      if np.isnan(nodata):
        band[band == 0] = aux
        band[np.isnan(band)] = 0
      else:
        band[band == 0] = aux
        band[band == nodata] = 0

    S = csr_matrix(band)
    list_indices = list(zip(S.nonzero()[0], S.nonzero()[1]))
    S = None

    band[band == 0] = no_data
    band[band == aux] = 0
    return band, nodata, list_indices
  else:
    return band, nodata

#Funcao que calcula o AGB de vegetacao secundaria
def calcIdadeSec(idade: int, estado: str, max_agb: float, razao_ini: float, a1: float) -> float:
  """
    Funcao que calcula o AGB de vegetacao secundaria.
    
    Entrada:
        idade: Numero inteiro da idade do pixel.
        estado: Nome da area de interesse.
        max_agb: Valor maximo de AGB da vegetacao secundaria.
        razao_ini: Valor da razao que descreve o crescimento de AGB por idade.
        a1: Valor inicial de AGB que recebera pixels com idade 1.
        
    Saída:
        result: Resultado do calculo do AGB convertido para carbono.
  """ 
  if 'Canada' in estado:
    result = 23.2*np.log(idade)-35.7
  else:
    a1 = a1
    razao = razao_ini
    if estado in ['PA','MA']:
      if idade >= 21:
        a1 = (a1 + (20 - 1)*razao)
        razao = (a1-max_agb)/(21-80)
        idade = idade - 19
    else:
      if idade >= 21:
        a1 = (a1 + (20 - 1)*razao)
        razao = (a1-max_agb)/(21-100)
        if estado == 'Malasia' or estado == "Indonesia":
          razao = (a1-max_agb)/(21-40)
        elif estado == 'Mocambique':
          razao = (a1-max_agb)/(20-100)
        idade = idade - 19
  
    result = a1 + (idade - 1)*razao
  
  if result < 0:
    result = 0
  return result*0.47#converte para carbono

#Funcao que verifica o crs
def checkCRS(raster_path: str, crs_std: str = 'EPSG:4326') -> bool:
  """
    Funcao que verfica a projecao de um raster.
    
    Entrada:
        raster_path: Caminho do raster cuja projecao sera verificada.
        crs_std: Projecao de interesse.
        
    Saída:
        Retorna um valor booleano True se o raster estiver na projecao correta, False o contrario.
  """ 
  with rasterio.open(raster_path) as dataset:
    crs = dataset.crs
  dataset = None

  if crs != crs_std:
    return False
  else:
    return True

#Funcao que ira juntar os pixels do Mapbiomas e IBGE
def join_ibge_mapbio(id_class: int, dict_classes: dict, dict_ibge: dict,
                     agb_ibge: dict, agb_p_estado: dict, est_interesse: str) -> Tuple[dict,dict]:
  """
    Funcao que ira juntar os pixels do Mapbiomas e IBGE.
    
    Entrada:
        id_class: Classe de vegetacao do tipo floresta.
        dict_classes: Dicionario com as classes de vegetacao principal.
        dict_ibge: Dicionario com as classes de vegetacao auxiliar.
        agb_ibge: Dicionario com os valores de AGB para as classes de vegetacao auxiliar.
        agb_p_estado: Dicionario com os valores de AGB para as classes de vegetacao principal.
        est_interesse: Area de interesse.
        
    Saída:
        dict_classes: Dicionario com as classes de vegetacao principal e auxiliar.
        agb_p_estado: Dicionario com os valores de AGB das classes principal e auxiliar.
  """ 
  #Junta os valores de AGB do Mapbiomas e IBGE separado por classe
  for est in agb_p_estado:
    for ibge_id in agb_ibge[est]:
      agb_p_estado[est]['natural'][ibge_id+300] = agb_ibge[est][ibge_id]

  lista_ibge_classes = list(set(list(dict_ibge.keys()))-set(list(agb_ibge[est_interesse].keys())))
  for i in lista_ibge_classes:
    del dict_ibge[i]

  for ibge_id in dict_ibge:
      dict_classes['natural'][id_class] = list(set(dict_classes['natural'][id_class])-set(dict_ibge[ibge_id]))

  for ibge_id in dict_ibge:
    dict_classes['natural'][ibge_id+300] = dict_ibge[ibge_id]

  return dict_classes, agb_p_estado

#Funcao que verifica se uma matriz possui nodata NaN
def check_is_nan(band: np.ndarray, nodata: Union[int, float, str, None]) -> Tuple[np.ndarray,float]:
  """
    Funcao que verifica se uma matriz possui nodata NaN.
    
    Entrada:
        band: Matriz numpy de pixels.
        nodata: Valor nodata da matriz.
        
    Saída:
        band: Matriz numpy de pixels atualizada.
        nodata: Valor nodata da matriz atualizado.
  """ 
  if np.isnan(nodata):
    band[np.isnan(band)] = -9999
    nodata = -9999
  else:
    band[band == nodata] = -9999
    nodata = -9999
  
  return band, nodata

#Funcao que cria o raster final
def make_raster(vetor_pixels: List[Tuple[int,int]], vetor_value: List[Tuple[float]], raster_std: str, est_interesse: str, 
                ano_interesse: int, tmp_path: str, name: str, final_name: Union[str,None], nodata: Union[int, float, str, None], 
                dtype: str, descriptions: str) -> str:
  """
    Funcao que cria o raster final.
    
    Entrada:
        vetor_pixels: Lista com os indices dos pixels.
        vetor_value: Lista com o valor dos pixels.
        raster_std: Caminho do raster que sera usado como padrao.
        est_interesse: Area de interesse.
        ano_interesse: Ano de interesse.
        tmp_path: Caminho da pasta onde sera salvo o raster gerado.
        name: Nome intermediario que estara no nome final do arquivo raster a ser gerado.
        final_name: Nome final do arquivo raster a ser gerado.
        nodata: Valor que os pixels nodata possuem.
        dtype: Variavel que indica qual o tipo de dados do raster (Ex: uint8).
        descriptions: String com descricao do raster que sera gerado.
        
    Saída:
        Retorna o caminho onde o raster gerado foi salvo.
  """ 
  with rasterio.open(raster_std) as dataset:
      out_meta = dataset.profile

      data = dataset.read(1)
      data = np.zeros(data.shape)
      if nodata == None:
        data[:] = np.nan
      else:
        data[:] = nodata

      if dtype == None:
        dtype = "uint8"

      for i in range(len(vetor_pixels)):
        data[vetor_pixels[i][0]][vetor_pixels[i][1]] = vetor_value[i]

      if nodata == None:
        out_meta.update(dtype=dtype,compress='lzw',nodata=np.nan)
      else:
        out_meta.update(dtype=dtype,compress='lzw',nodata=nodata)
      if final_name == None:
        final_name = tmp_path+"raster_"+name+"_"+est_interesse+"_"+str(ano_interesse)+".tif"
      else:
        final_name = tmp_path+final_name
  #Salva o raster de AGB que foi criado
      with rasterio.open(final_name, "w", **out_meta) as dest:
          dest.set_band_description(1, descriptions)
          dest.write(data, indexes=1)

  dataset = None
  dest = None
  data = None

  return final_name

#Funcao que gera o vetor de pixels e AGB
def make_raster_indices(dict_classes: dict, dict_agb: dict, agb_p_estado: dict, est_interesse: str,
                        max_agb: float, razao_ini: float, a1: float) -> Tuple[List,List]:
  """
    Funcao que gera o vetor de pixels e AGB.
    
    Entrada:
        dict_classes: Dicionario com as classes de vegetacao principal.
        dict_agb: Dicionario com os valores de AGB para as classes de vegetacao.
        agb_p_estado: Dicionario com os valores de AGB para as classes de vegetacao principal.
        est_interesse: Area de interesse.
        max_agb: Valor maximo de AGB da vegetacao secundaria.
        razao_ini: Valor da razao que descreve o crescimento de AGB por idade.
        a1: Valor inicial de AGB que recebera pixels com idade 1.
        
    Saída:
        vetor_pixels: Lista com os indices onde estarao os pixels.
        vetor_value: Lista com o valor de cada pixel.
  """ 
  #Cria vetor de pixels e de valor de AGB
  vetor_pixels = []
  vetor_value = []

  #Adiciona os pixels e agb das florestas naturais
  for veg in dict_classes['natural']:
    for p in range(len(dict_classes['natural'][veg])):
      vetor_pixels.append(dict_classes['natural'][veg][p])
      vetor_value.append(dict_agb[veg][p])

  #Adiciona os pixels e agb da vegetacao secundaria
  for idade in dict_classes['secundaria']:
    AGB_value = calcIdadeSec(idade,est_interesse,max_agb,razao_ini,a1)
    for p in range(len(dict_classes['secundaria'][idade])):
      vetor_pixels.append(dict_classes['secundaria'][idade][p])
      vetor_value.append(AGB_value)

  #Adiciona os pixels e agb da vegetacao antropica
  for forest in dict_classes['antropica']:
    for p in range(len(dict_classes['antropica'][forest])):
      vetor_pixels.append(dict_classes['antropica'][forest][p])
      vetor_value.append(agb_p_estado[est_interesse]['antropica'][forest])
  
  return vetor_pixels, vetor_value

#Funcao que copia um arquivo para o diretorio de saida
def copy_raster(file_path: str, out_path: str):
  """
    Funcao que copia um arquivo para o diretorio de saida.
    
    Entrada:
        file_path: Caminho do arquivo que sera copiado.
        out_path: Caminho para onde o arquivo sera copiado.
        
    Saída:
  """
  if not os.path.exists(out_path):
    os.makedirs(out_path)
  command = "cp "+file_path+" "+out_path+""
  os.system(command)

#Funcao que exclui um diretorio
def del_folder(dest: str):
  """
    Funcao que deleta um diretorio.
    
    Entrada:
        dest: Caminho do diretorio que sera deletado.

    Saída:
  """
  import shutil
  shutil.rmtree(dest, ignore_errors=True)

#Funcao que calcula o sha256 de um arquivo e salva em um tx
def checksum_sha256(file_path: str) -> str:
  """
    Funcao que calcula o sha256 de um arquivo e salva em um txt.
    
    Entrada:
        file_path: Caminho do arquivo que sera calculado o sha256.        
    
    Saída:
        Retorna o caminho onde o arqivo txt foi criado.
  """
  import hashlib
  
  sha256_hash = hashlib.sha256()
  with open(file_path,"rb") as f:
    # Read and update hash string value in blocks of 4K
    for byte_block in iter(lambda: f.read(4096),b""):
        sha256_hash.update(byte_block)
    file_hash = sha256_hash.hexdigest()
  f.close()

  name = file_path.split('.')[0]

  with open(name+"_sha256.txt",mode="w",encoding="utf-8") as archive:
    archive.write(file_hash)
  archive.close

  return name+"_sha256.txt"

#Funcao que "zipa" um arquivo
def zip_file(path_file: str, path_output: str, name: str) -> str:
  """
    Funcao que comprime um arquivo em zip.
    
    Entrada:
        file_path: Caminho do arquivo que sera coprimido.
        path_output: Caminho onde sera criado o arquivo zip.
        name: nome final do arquivo zip.         
    
    Saída:
        Retorna o caminho onde o arqivo zip foi criado.
  """

  import zipfile
  try:
      import zlib
      compression = zipfile.ZIP_DEFLATED
  except:
      compression = zipfile.ZIP_STORED

  modes = { zipfile.ZIP_DEFLATED: 'deflated',
            zipfile.ZIP_STORED:   'stored',
            }
  if name == None:
    name = 'zip_file.zip'
  else:
    if '.zip' not in name:
      name += '.zip'

  zf = zipfile.ZipFile(path_output+name, mode='w')
  try:
      zf.write(path_file, path_file.split('/')[-1],compress_type=compression)
  finally:
      zf.close()

  return path_output+name

#Funcao que adiciona um arquivo em um zip preexistente
def add_zip(path_zip: str, path_file: str):
  """
    Funcao que adiciona um arquivo em um zip preexistente.
    
    Entrada:
        path_zip: Caminho do arquivo zip.
        path_file: Caminho do arquivo que sera adicionado ao zip.
    
    Saída:
  """
  import zipfile
  try:
      import zlib
      compression = zipfile.ZIP_DEFLATED
  except:
      compression = zipfile.ZIP_STORED

  modes = { zipfile.ZIP_DEFLATED: 'deflated',
            zipfile.ZIP_STORED:   'stored',
            }

  z = zipfile.ZipFile(path_zip, mode="a")
  try:
      z.write(path_file, path_file.split('/')[-1],compress_type=compression)
  finally:
      z.close()

def create_dir(dir_path: str) -> str:
  """
    Funcao que cria uma pasta localmente.
    
    Entrada:
        dir_path: String contendo caminho onde sera criada a pasta.
    
    Saída: String contendo o caminho onde a pasta foi criada.
  """
  if not os.path.exists(dir_path):
    os.makedirs(dir_path)
  dir_path = dir_path+'/'
  return dir_path

def create_raster_dict(band: np.ndarray, list_indices: List[Tuple]) -> dict:
  """
    Funcao que cria um dicionario de classes de um mapa. Onde cada chave do dicionario
    sera uma classe do mapa que contera uma lista com os indices dos pixels valorados.
    
    Entrada:
        band: Array Numpy do arquivo raster.
        list_indices: Lista dos pixels valorados que existem no mapa.
    
    Saída: Dicionario de pixels valorados por classe do pixel.
  """
  dict_classes = {}
  for ind in list_indices:
    pixel = band[ind[0]][ind[1]]
    if pixel not in dict_classes:
      dict_classes[pixel] = []
      dict_classes[pixel].append(ind)
    else:
      dict_classes[pixel].append(ind)
  return dict_classes

def extract_file(archive: str, file_name: str, zip_path: str) -> str:
  """
    Funcao que extrai um arquivo especifico de dentro de um arquivo zip.
    
    Entrada:
        archive: String contendo o caminho do arquivo zip.
        file_name: String contendo o nome do arquivo que se deseja extrair.
        zip_path: String contendo a pasta onde sera extraido o arquivo.
    
    Saída: String do caminho onde foi extraido o arquivo.
  """
  
  arc = zipfile.ZipFile(archive, 'r')
  arc.extract(file_name, zip_path)
  
  return zip_path+"/"+file_name

def get_raster_band(raster_path: str) -> Tuple[np.ndarray, Union[int, float, str, None]]:
  """
      Funcao que acessa um raster e retorna a sua banda.
      
      Entrada:
          raster_path: String contendo o caminho do arquivo raster.
      
      Saída:
          band: Matriz numpy contendo a banda 1 do raster.
          nodata: Variavel contendo o valor de pixel nodata do raster.
    """

  with rasterio.open(raster_path) as dataset:
    band = dataset.read(1)
    nodata = dataset.nodata
  dataset.close()
  dataset = None

  return band, nodata

def get_shapes_from_file(shp_path: str) -> dict:
  """
    Funcao que obtem os shapes existentes em um shapefile.
    
    Entrada:
        shp_path: String contendo o caminho do arquivo shapefile.
    
    Saída:
        shapes_dict: Dicionario contendo os shapes do arquivo shapefile.
  """

  shapes_dict = {}
  with fiona.open(shp_path, "r") as shapefile:
    for c,feature in enumerate(shapefile):
      shapes_dict[c] = [feature["geometry"]]

  del shapefile
  return shapes_dict

def calculate_area_from_shape(shp_path: str) -> pd.DataFrame:
  """
    Funcao que calcula a area de um shapefile.
    
    Entrada:
        shp_path: String contendo o caminho do arquivo shapefile.
    
    Saída:
        df_tmp: DataFrame contendo a area do shapefile.
  """
  data = gpd.read_file(shp_path)
  data = data.to_crs({'proj':'cea'})
  df_tmp = pd.DataFrame()
  df_tmp['area'] = data['geometry'].area#/ 10**6 / 100
  # dividir por 10**6 retorna area em km2

  return df_tmp

def fractional_pixel_weights(raster_path: str, geom: Any, list_pixels: List[Tuple[int,int]]) -> np.ndarray:
  """
    Funcao que calcula a fracao de area de um shapefile que intersecta com um pixel, 
    i.e. percentual de quanto cada pixel é coberto pela area do shape.
    
    Entrada:
        raster_path: String contendo o caminho do arquivo raster.
        geom: Objeto do tipo geometry que representa o shapefile.
        list_pixels: Lista de tuplas contendo os indices dos pixels.
    
    Saída:
        frac_intersected: Array numpy contendo o percentual de area coberta por pixel.
  """
  
  geom = shape(geom)
  fsrc = rasterio.open(raster_path)

  gt = fsrc.profile['transform']
  xs = np.arange(gt[2], gt[2] +  gt[0]* (1 + fsrc.shape[1]), gt[0])
  ys = np.arange(gt[5], gt[5] +  gt[4]* (1 + fsrc.shape[0]), gt[4])

  # Convert geom into ogr geometry
  geom_ogr = ogr.CreateGeometryFromWkt(geom.wkt)

  # Loop through each grid cell, compute the intersecting area
  overlapping_areas = np.empty((len(ys)-1, len(xs)-1))

  for ind in list_pixels:
    ix = ind[1]
    iy = ind[0]

    xmin = xs[ix]
    xmax = xs[ix + 1]
    ymax = ys[iy]
    ymin = ys[iy + 1]

    # Intersecting area
    coords_wkt = "POLYGON ((" + str(xmin) + ' ' + str(ymax) + ', ' + str(xmax) + ' ' + str(ymax) + ', ' + str(xmax) + ' ' + str(ymin) + ', ' + str(xmin) + ' ' + str(ymin) + ', ' + str(xmin) + ' ' + str(ymax) + "))"
    polycell = ogr.CreateGeometryFromWkt(coords_wkt)
    overlapping_areas[iy, ix] = polycell.Intersection(geom_ogr).Area()

  # Ratio of overlapped area to pixel area
  frac_intersected = overlapping_areas / (abs(gt[0] * gt[4]))

  return frac_intersected